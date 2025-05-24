#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Módulo para extraer datos de un archivo Excel
"""
import openpyxl

# Necesitamos añadir o corregir el método extract_historic_data en la clase ExcelDataExtractor

class ExcelDataExtractor:
    """
    Clase para extraer datos de un archivo Excel
    """
    
    def __init__(self, excel_path):
        """
        Inicializa el extractor con la ruta al archivo Excel
        
        Args:
            excel_path (str): Ruta al archivo Excel
        """
        self.excel_path = excel_path
        # Cargar el workbook aquí para evitar el error
        try:
            print(f"Cargando archivo Excel: {excel_path}")
            self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
            print(f"Hojas disponibles: {self.workbook.sheetnames}")
        except Exception as e:
            print(f"Error al cargar el archivo Excel: {str(e)}")
            raise
    
    def extract_data(self, sheet_name):
        """
        Extrae datos de una hoja específica del Excel
        
        Args:
            sheet_name (str): Nombre de la hoja a extraer
            
        Returns:
            list: Lista de diccionarios con los datos extraídos
        """
        try:
            # Verificar que la hoja existe
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo Excel")
                
            # Cargar la hoja de Excel
            sheet = self.workbook[sheet_name]
            
            # Obtener encabezados (primera fila)
            headers = []
            for cell in sheet[1]:
                header_value = cell.value
                if header_value is not None:
                    header_value = str(header_value).strip()
                headers.append(header_value)
            
            # Imprimir encabezados para depuración
            print(f"Encabezados encontrados en {sheet_name}: {headers}")
            
            # Extraer datos
            data = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                # Crear un diccionario con los datos de la fila
                row_data = {}
                has_data = False
                
                # Cargar todos los datos originales
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx] is not None:
                        row_data[headers[col_idx]] = value
                        # Verificar si hay algún valor no nulo
                        if value is not None:
                            has_data = True
                
                # Solo añadir filas que tengan algún dato
                if has_data:
                    # Procesar campos de jerarquía
                    if 'CIA' in row_data:
                        row_data['company'] = str(row_data['CIA']).strip() if row_data['CIA'] is not None else ""
                        print(f"Fila {row_idx}: CIA = '{row_data['company']}'")
                    
                    if 'PRJID' in row_data:
                        row_data['project'] = str(row_data['PRJID']).strip() if row_data['PRJID'] is not None else ""
                        print(f"Fila {row_idx}: PRJID = '{row_data['project']}'")
                    
                    if 'ROW' in row_data:
                        row_data['Category'] = str(row_data['ROW']).strip() if row_data['ROW'] is not None else "Sin categoría"
                        print(f"Fila {row_idx}: ROW (Category) = '{row_data['Category']}'")
                    
                    if 'COLUMN' in row_data:
                        row_data['Subcategory'] = str(row_data['COLUMN']).strip() if row_data['COLUMN'] is not None else "Sin subcategoría"
                        print(f"Fila {row_idx}: COLUMN (Subcategory) = '{row_data['Subcategory']}'")
                    
                    # Datos específicos para históricos: línea temporal (formato YYYY.WW)
                    if 'WKS' in row_data and row_data['WKS'] is not None:
                        try:
                            # Convertir a string y formatear correctamente
                            period_str = str(row_data['WKS']).strip()
                            # Verificar si tiene el formato correcto YYYY.WW
                            if '.' in period_str:
                                year_part, week_part = period_str.split('.')
                                # Asegurar que la semana tenga 2 dígitos
                                if len(week_part) == 1:
                                    week_part = '0' + week_part
                                period_str = f"{year_part}.{week_part}"
                            row_data['period'] = period_str
                            print(f"Fila {row_idx}: Periodo (WKS) = '{row_data['period']}'")
                        except Exception as e:
                            print(f"Error al procesar periodo en fila {row_idx}: {e}")
                            row_data['period'] = str(row_data['WKS'])
                    
                    # Valores monetarios para series temporales
                    if 'PREV' in row_data:
                        try:
                            row_data['prev'] = float(row_data['PREV']) if row_data['PREV'] is not None else 0.0
                            print(f"Fila {row_idx}: PREV = {row_data['prev']}")
                        except (ValueError, TypeError):
                            row_data['prev'] = 0.0
                            print(f"Fila {row_idx}: PREV = {row_data['prev']} (convertido de valor no numérico)")
                    
                    if 'PPTO' in row_data:
                        try:
                            row_data['ppto'] = float(row_data['PPTO']) if row_data['PPTO'] is not None else 0.0
                            print(f"Fila {row_idx}: PPTO = {row_data['ppto']}")
                        except (ValueError, TypeError):
                            row_data['ppto'] = 0.0
                            print(f"Fila {row_idx}: PPTO = {row_data['ppto']} (convertido de valor no numérico)")
                    
                    # Para datos de KPI, mapear campos específicos
                    # PREV ya está procesado arriba
                    if 'PREV' in row_data and row_data['PREV'] is not None:
                        try:
                            row_data['prevValue'] = float(row_data['PREV'])
                            print(f"Fila {row_idx}: PREV = {row_data['prevValue']} (mapeado a prevValue)")
                        except (ValueError, TypeError):
                            row_data['prevValue'] = 0.0
                            print(f"Fila {row_idx}: PREV = {row_data['prevValue']} (convertido de valor no numérico)")
                    
                    if 'PDTE' in row_data and row_data['PDTE'] is not None:
                        try:
                            row_data['pendingValue'] = float(row_data['PDTE'])
                            print(f"Fila {row_idx}: PDTE = {row_data['pendingValue']} (mapeado a pendingValue)")
                        except (ValueError, TypeError):
                            row_data['pendingValue'] = 0.0
                            print(f"Fila {row_idx}: PDTE = {row_data['pendingValue']} (convertido de valor no numérico)")
                    
                    if 'REALPREV' in row_data and row_data['REALPREV'] is not None:
                        try:
                            row_data['realPrevPercent'] = float(row_data['REALPREV'])
                            print(f"Fila {row_idx}: REALPREV = {row_data['realPrevPercent']} (mapeado a realPrevPercent)")
                        except (ValueError, TypeError):
                            row_data['realPrevPercent'] = 0.0
                            print(f"Fila {row_idx}: REALPREV = {row_data['realPrevPercent']} (convertido de valor no numérico)")
                    
                    if 'PPTOPREV' in row_data and row_data['PPTOPREV'] is not None:
                        try:
                            row_data['pptoPrevPercent'] = float(row_data['PPTOPREV'])
                            print(f"Fila {row_idx}: PPTOPREV = {row_data['pptoPrevPercent']} (mapeado a pptoPrevPercent)")
                        except (ValueError, TypeError):
                            row_data['pptoPrevPercent'] = 0.0
                            print(f"Fila {row_idx}: PPTOPREV = {row_data['pptoPrevPercent']} (convertido de valor no numérico)")
                    
                    # Generar una clave única basada en la jerarquía
                    hierarchy_key = []
                    for field in ['company', 'project', 'Category', 'Subcategory']:
                        if field in row_data and row_data[field]:
                            hierarchy_key.append(str(row_data[field]))
                    
                    if hierarchy_key:
                        row_data['hierarchyKey'] = '|'.join(hierarchy_key)
                        print(f"Fila {row_idx}: Clave jerárquica = '{row_data['hierarchyKey']}'")
                    
                    # Preparar datos para series temporales si es un registro histórico
                    if 'period' in row_data:
                        row_data['isHistorical'] = True
                        # Crear estructura para series temporales
                        row_data['timeSeriesData'] = {
                            'period': row_data['period'],
                            'real': row_data.get('real', 0.0),
                            'prev': row_data.get('prev', 0.0),
                            'ppto': row_data.get('ppto', 0.0)
                        }
                        print(f"Fila {row_idx}: Datos de serie temporal creados: {row_data['timeSeriesData']}")
                    
                    data.append(row_data)
            
            print(f"Total de filas extraídas de {sheet_name}: {len(data)}")
            
            # Imprimir ejemplo de datos para depuración con más campos
            if data:
                print(f"Ejemplo de datos extraídos de {sheet_name}:")
                example_data = data[0]
                # Mostrar todos los campos importantes
                important_fields = ['PRJID', 'CIA', 'ROW', 'COLUMN', 'WKS', 'PREV', 'PPTO', 
                                   'PDTE', 'REALPREV', 'PPTOPREV']
                
                print("  --- Campos originales ---")
                for field in important_fields:
                    if field in example_data:
                        print(f"  {field}: {example_data[field]}")
                
                # Mostrar campos derivados
                derived_fields = ['company', 'project', 'Category', 'Subcategory', 'period', 
                                 'prev', 'ppto', 'prevValue', 
                                 'pendingValue', 'realPrevPercent', 'pptoPrevPercent', 'hierarchyKey']
                
                print("  --- Campos derivados/mapeados ---")
                for field in derived_fields:
                    if field in example_data:
                        print(f"  {field}: {example_data[field]}")
            
            return data
            
        except Exception as e:
            print(f"Error al extraer datos de {sheet_name}: {str(e)}")
            raise
    
    def extract_kpi_data(self, sheet_name):
        """
        Extrae datos KPI de la hoja especificada
        """
        try:
            print(f"Extrayendo datos KPI de la hoja: {sheet_name}")
            
            # Definir los encabezados esperados para FrmBB_3
            expected_headers = ['CIA', 'PRJID', 'ROW', 'COLUMN', 'KPREV', 'PDTE', 'REALPREV', 'PPTOPREV']
            
            # Cargar la hoja especificada
            sheet = self.workbook[sheet_name]
            
            # Inicializar lista para almacenar los datos
            kpi_data = []
            
            # Obtener los valores de la primera fila (encabezados)
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            print(f"Encabezados encontrados en {sheet_name}: {header_row}")
            
            # Crear un mapeo de índice a nombre de columna esperado
            header_map = {}
            for idx, header in enumerate(header_row):
                if header is not None and idx < len(expected_headers):
                    header_map[idx] = expected_headers[idx]
            
            print(f"Mapeo de encabezados: {header_map}")
            
            # Procesar filas de datos (a partir de la segunda fila)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                # Omitir filas vacías
                if not any(row):
                    continue
                
                # Crear un diccionario con los datos de la fila
                row_data = {}
                for idx, value in enumerate(row):
                    if idx in header_map:
                        row_data[header_map[idx]] = value
                
                # Asegurarse de que tenemos ROW y COLUMN
                if 'ROW' in row_data and 'COLUMN' in row_data and row_data['ROW'] is not None and row_data['COLUMN'] is not None:
                    kpi_data.append(row_data)
            
            print(f"Total de filas KPI extraídas: {len(kpi_data)}")
            
            # Mostrar ejemplo de datos para depuración
            if kpi_data and len(kpi_data) > 0:
                print("Ejemplo de datos KPI:")
                print(kpi_data[0])
            
            return kpi_data
            
        except Exception as e:
            import traceback
            print(f"Error al extraer datos KPI: {e}")
            traceback.print_exc()
            return []
    
    def extract_historic_data(self, sheet_name):
        """
        Extrae datos históricos de la hoja especificada
        """
        try:
            print(f"Extrayendo datos históricos de la hoja: {sheet_name}")
            
            # Definir los encabezados esperados para FrmBB_2
            expected_headers = ['CIA', 'PRJID', 'ROW', 'COLUMN', 'WKS', 'REAL', 'PPTO', 'HPREV']
            
            # Cargar la hoja especificada
            sheet = self.workbook[sheet_name]
            
            # Inicializar lista para almacenar los datos
            historic_data = []
            
            # Obtener los valores de la primera fila (encabezados)
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            print(f"Encabezados encontrados en {sheet_name}: {header_row}")
            
            # Crear un mapeo de índice a nombre de columna esperado
            header_map = {}
            for idx, header in enumerate(header_row):
                if header is not None and idx < len(expected_headers):
                    header_map[idx] = expected_headers[idx]
            
            print(f"Mapeo de encabezados: {header_map}")
            
            # Procesar filas de datos (a partir de la segunda fila)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                # Omitir filas vacías
                if not any(row):
                    continue
                
                # Crear un diccionario con los datos de la fila
                row_data = {}
                for idx, value in enumerate(row):
                    if idx in header_map:
                        row_data[header_map[idx]] = value
                
                # Asegurarse de que tenemos ROW, COLUMN y WKS (que es PERIOD)
                if ('ROW' in row_data and row_data['ROW'] is not None and 
                    'COLUMN' in row_data and row_data['COLUMN'] is not None and 
                    'WKS' in row_data and row_data['WKS'] is not None):
                    
                    # Añadir PERIOD como alias de WKS para mantener compatibilidad
                    row_data['PERIOD'] = row_data['WKS']
                    historic_data.append(row_data)
            
            print(f"Total de filas históricas extraídas: {len(historic_data)}")
            
            # Mostrar ejemplo de datos para depuración
            if historic_data and len(historic_data) > 0:
                print("Ejemplo de datos históricos:")
                print(historic_data[0])
            
            return historic_data
            
        except KeyError as e:
            print(f"Error: La hoja '{sheet_name}' no existe en el archivo Excel.")
            return None
        except Exception as e:
            import traceback
            print(f"Error al extraer datos históricos: {e}")
            traceback.print_exc()
            return None