#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Módulo para extraer datos de un archivo Excel
"""
import openpyxl

# Necesitamos añadir o corregir el método extract_historic_data en la clase ExcelDataExtractor

class ExcelDataExtractor:
    """
    Clase para extraer datos de un archivo Excel
    """
    
    def __init__(self, excel_path):
        """
        Inicializa el extractor con la ruta al archivo Excel
        """
        self.excel_path = excel_path
        try:
            print(f"Cargando archivo Excel: {excel_path}")
            self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
            print(f"Hojas disponibles: {self.workbook.sheetnames}")
        except Exception as e:
            print(f"Error al cargar el archivo Excel: {str(e)}")
            raise
    
    def extract_data(self, kpi_sheet_name, historic_sheet_name):
        """
        Coordina la extracción de datos KPI e históricos.

        Args:
            kpi_sheet_name (str): Nombre de la hoja de datos KPI.
            historic_sheet_name (str): Nombre de la hoja de datos históricos.

        Returns:
            dict: Diccionario con los datos extraídos de ambas hojas.
        """
        print(f"Extrayendo datos KPI desde la hoja: {kpi_sheet_name}")
        kpi_data = self.extract_kpi_data(kpi_sheet_name)

        print(f"Extrayendo datos históricos desde la hoja: {historic_sheet_name}")
        historic_data = self.extract_historic_data(historic_sheet_name)

        data = {
            "kpiData": kpi_data,
            "historicData": historic_data
        }
        return data
    
    def extract_kpi_data(self, sheet_name):
        """
        Extrae datos KPI de la hoja especificada
        """
        try:
            print(f"Extrayendo datos KPI de la hoja: {sheet_name}")

            # Definir los encabezados esperados para FrmBB_3
            expected_headers = ['CIA', 'PRJID', 'ROW', 'COLUMN', 'KPREV', 'PDTE', 'REALPREV', 'PPTOPREV']

            # Cargar la hoja especificada
            if sheet_name not in self.workbook.sheetnames:
                print(f"Error: La hoja '{sheet_name}' no existe en el archivo Excel.")
                return []

            sheet = self.workbook[sheet_name]

            # Inicializar lista para almacenar los datos
            kpi_data = []

            # Obtener los valores de la primera fila (encabezados)
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

            # Filtrar encabezados válidos y eliminar valores None
            header_row = [header for header in header_row if header is not None]
            print(f"Encabezados encontrados en {sheet_name}: {header_row}")

            # Crear un mapeo de índice a nombre de columna esperado
            header_map = {}
            for idx, header in enumerate(header_row):
                if header is not None and header in expected_headers:
                    header_map[idx] = header

            print(f"Mapeo de encabezados: {header_map}")

            # Procesar filas de datos (a partir de la segunda fila)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                # Omitir filas vacías
                if not any(row):
                    continue

                # Crear un diccionario con los datos de la fila
                row_data = {}
                for idx, value in enumerate(row):
                    if idx in header_map:
                        row_data[header_map[idx]] = value

                # Asegurarse de que tenemos ROW y COLUMN
                if 'ROW' in row_data and 'COLUMN' in row_data and row_data['ROW'] is not None and row_data['COLUMN'] is not None:
                    kpi_data.append(row_data)

            print(f"Total de filas KPI extraídas: {len(kpi_data)}")

            # Mostrar ejemplo de datos para depuración
            if kpi_data and len(kpi_data) > 0:
                print("Ejemplo de datos KPI:")
                print(kpi_data[0])

            return kpi_data

        except Exception as e:
            import traceback
            print(f"Error al extraer datos KPI: {e}")
            traceback.print_exc()
            return []
    
    def normalize(self, val):
        """
        Normaliza un valor eliminando espacios y convirtiéndolo a mayúsculas.
        """
        if val is None:
            return ""
        return str(val).strip().upper()

    def extract_historic_data(self, sheet_name):
        """
        Extrae datos históricos de una hoja específica del Excel.
        """
        try:
            print(f"Extrayendo datos históricos de la hoja: {sheet_name}")
            sheet = self.workbook[sheet_name]
            data_rows = list(sheet.rows)[1:]
            headers = [cell.value for cell in list(sheet.rows)[0]]

            # Filtrar encabezados válidos y eliminar valores None
            headers = [header for header in headers if header is not None]
            print(f"Encabezados encontrados: {headers}")

            hierarchy_keys = ['CIA', 'PRJID', 'ROW', 'COLUMN']
            value_keys = ['HPREV', 'PPTO', 'REAL']
            period_key = 'WKS'

            def clean_number(val):
                if val is None:
                    return 0.0
                if isinstance(val, (int, float)):
                    return float(val)
                s = str(val).replace("€", "").replace(" ", "").replace(".", "").replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return 0.0

            historic_data = []
            for row in data_rows:
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers) and headers[i] is not None:
                        col = headers[i]
                        if col in hierarchy_keys or col in value_keys or col == period_key:
                            row_data[col] = cell.value

                for key in hierarchy_keys:
                    row_data[key] = self.normalize(row_data.get(key))

                # Eliminar la asignación de PERIOD y usar WKS directamente
                if all(row_data.get(k, "") != "" for k in hierarchy_keys + [period_key]):
                    for key in value_keys:
                        if key in row_data:
                            row_data[key] = clean_number(row_data[key])
                        else:
                            row_data[key] = 0.0
                    historic_data.append(row_data)

            # Agrupar todos los registros de la serie temporal WKS para una celda jerárquica
            if historic_data:
                grouped_data = {}
                for record in historic_data:
                    hierarchy_key = tuple(record[key] for key in ['CIA', 'PRJID', 'ROW', 'COLUMN'])
                    if hierarchy_key not in grouped_data:
                        grouped_data[hierarchy_key] = []
                    grouped_data[hierarchy_key].append({
                        'WKS': record['WKS'],
                        'REAL': record['REAL'],
                        'PPTO': record['PPTO'],
                        'HPREV': record['HPREV']
                    })

                # Mostrar ejemplo de datos históricos agrupados
                example_key = next(iter(grouped_data))
                print("Ejemplo de datos históricos agrupados:")
                print({
                    'Hierarchy': dict(zip(['CIA', 'PRJID', 'ROW', 'COLUMN'], example_key)),
                    'Series': grouped_data[example_key]
                })

            return historic_data
        except Exception as e:
            print(f"Error al extraer datos históricos: {e}")
            import traceback
            traceback.print_exc()
            return []
